from flask import Flask, request, redirect, url_for, session, render_template, jsonify, send_file
import qrcode
import os
import os.path
import socket
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import json
from collections import defaultdict
import numpy as np
from flask_sqlalchemy import SQLAlchemy
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import re
import PyPDF2
import csv

app = Flask(__name__)
app.secret_key = "sai123"  # For session management

# Database Configuration
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///hackthon.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# Data storage (demo purpose)
customer_data = {}
user_history = defaultdict(list)  # Track user mood history
admin_overrides = {}  # Store admin recommendation overrides
ADMIN_USER = "sai"
ADMIN_PASS = "sai@143"

# ---------- DATABASE MODELS ----------
class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    age = db.Column(db.Integer, nullable=True)
    mobile = db.Column(db.String(20), nullable=True)
    email = db.Column(db.String(100), nullable=True)
    pdf_filename = db.Column(db.String(255), nullable=True)  # Store uploaded PDF filename
    pdf_uploaded_at = db.Column(db.DateTime, nullable=True)  # Track when PDF was uploaded
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    
    def __repr__(self):
        return f'<Customer {self.name}>'

class Feedback(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.String(50), nullable=False)
    rating = db.Column(db.Integer, nullable=False)  # 1-5 rating
    suggestion = db.Column(db.Text, nullable=True)  # Text feedback
    timestamp = db.Column(db.DateTime, default=datetime.now)
    status = db.Column(db.String(20), default='unread')  # unread/read
    
    def __repr__(self):
        return f'<Feedback {self.customer_id}>'

class FoodItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    base_price = db.Column(db.Float, nullable=False)
    quantity = db.Column(db.Integer, default=0)
    image_url = db.Column(db.String(200), nullable=True)
    is_available = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    
    def __repr__(self):
        return f'<FoodItem {self.name}>'

class Transaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.String(50), nullable=False)
    food_item_id = db.Column(db.Integer, db.ForeignKey('food_item.id'), nullable=False)
    quantity_purchased = db.Column(db.Integer, nullable=False)
    price_paid = db.Column(db.Float, nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.now)
    
    def __repr__(self):
        return f'<Transaction {self.customer_id} - {self.id}>'

class Mood(db.Model):
    """Store customer mood tracking data"""
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.String(50), nullable=False)
    mood = db.Column(db.String(50), nullable=False)  # e.g., Happy, Sad, Neutral
    intensity = db.Column(db.Integer, default=3)    # 1-5 scale
    notes = db.Column(db.Text, nullable=True)       # Optional personal notes
    timestamp = db.Column(db.DateTime, default=datetime.now)
    
    def __repr__(self):
        return f'<Mood {self.customer_id} - {self.mood}>'

# Emotion intensity levels
EMOTION_LEVELS = {
    "Very Happy": 5,
    "Happy": 4,
    "Neutral": 3,
    "Sad": 2,
    "Very Sad": 1,
    "Stressed": 2,
    "Calm": 4,
    "Excited": 5,
    "Tired": 2,
    "Energetic": 5,
    "Angry": 1,
    "Relaxed": 4,
    "Anxious": 1,
    "Focused": 4
}

# ---------- PDF EXTRACTION FUNCTION ----------
def extract_data_from_pdf(pdf_file):
    """Extract customer information from PDF file"""
    try:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        extracted_text = ""
        
        # Extract text from all pages
        for page in pdf_reader.pages:
            extracted_text += page.extract_text() + "\n"
        
        # Extract information using regex patterns
        data = {}
        
        # Try to extract name (looks for 'Name' or 'name' followed by colon)
        name_match = re.search(r'(?:Name|name)\s*[:]\s*([^\n]+)', extracted_text)
        if name_match:
            data['name'] = name_match.group(1).strip()
        
        # Try to extract age
        age_match = re.search(r'(?:Age|age)\s*[:]\s*(\d+)', extracted_text)
        if age_match:
            data['age'] = int(age_match.group(1))
        
        # Try to extract mobile/phone number
        mobile_match = re.search(r'(?:Mobile|mobile|Phone|phone)\s*[:]\s*([\d\s\-\+]+)', extracted_text)
        if mobile_match:
            data['mobile'] = mobile_match.group(1).strip()
        
        # Try to extract email
        email_match = re.search(r'(?:Email|email)\s*[:]\s*([^\s\n]+@[^\s\n]+)', extracted_text)
        if email_match:
            data['email'] = email_match.group(1).strip()
        
        return data
    except Exception as e:
        print(f"Error extracting PDF data: {e}")
        return {}

def save_uploaded_pdf(pdf_file, customer_id):
    """Save uploaded PDF file to uploads folder"""
    try:
        if not os.path.exists('static/uploads'):
            os.makedirs('static/uploads')
        
        # Create filename with customer ID and timestamp
        file_ext = '.pdf'
        filename = f"customer_{customer_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{file_ext}"
        filepath = os.path.join('static/uploads', filename)
        
        # Save the file
        pdf_file.save(filepath)
        return filename
    except Exception as e:
        print(f"Error saving PDF: {e}")
        return None


# ---------- AI SUGGESTION ENGINE WITH PERSONALIZATION ----------
def ai_suggestion(cid, mood):
    """Generate suggestions based on mood and user history"""
    base_suggestions = {
        "Very Happy": "🎶 Upbeat music | 💡 Bright lights | 🥗 Healthy snacks | 🎉 Celebrate moment",
        "Happy": "🎶 Upbeat music | 💡 Bright lights | 🥗 Healthy snacks",
        "Neutral": "🎵 Balanced music | 💭 Journaling | ☕ Coffee break",
        "Sad": "🎧 Comfort music | 🔆 Warm lights | 🍫 Chocolate | 👥 Talk to someone",
        "Very Sad": "🎧 Comfort music | 🔆 Warm lights | 🍫 Treat yourself | 🏥 Seek support",
        "Stressed": "🎼 Calm music | 🌙 Dim lights | ☕ Herbal tea | 🧘 Meditation",
        "Calm": "🌿 Nature sounds | 🌙 Soft lights | 🚶 Walk outside",
        "Excited": "🎯 Plan activities | 📱 Share news | 🎉 Celebrate | 🏃 Exercise",
        "Tired": "🎵 Soft music | 🌙 Soft lights | 🍌 Energy fruits | 😴 Rest",
        "Energetic": "💪 Exercise | 🏃 Sports | 🎯 Productive tasks",
        "Angry": "🌿 Nature sounds | ❄ Cool blue lights | 🥤 Fresh juice | 🧘 Breathing exercise",
        "Relaxed": "🎵 Soft music | tea ☕ | 📖 Reading",
        "Anxious": "🧘 Deep breathing | 🌿 Nature | 📱 Call friend | 🎵 Calming music",
        "Focused": "🎵 Focus music | 📚 Study space | ☕ Coffee"
    }
    
    suggestion = base_suggestions.get(mood, "Relaxing environment")
    
    # Add personalized touch based on history
    if cid in user_history and len(user_history[cid]) > 3:
        history_moods = [h['mood'] for h in user_history[cid][-5:]]
        if mood == history_moods[-1] and history_moods.count(mood) > 2:
            suggestion += " | ⚠️ Consistent mood - consider professional support"
    
    return suggestion

def get_personalized_recommendations(cid):
    """Get recommendations based on user history"""
    if cid not in user_history or len(user_history[cid]) == 0:
        return "No history yet. Keep sharing your mood!"
    
    moods_history = [h['mood'] for h in user_history[cid]]
    mood_freq = defaultdict(int)
    for mood in moods_history:
        mood_freq[mood] += 1
    
    most_common = max(mood_freq, key=mood_freq.get)
    return f"Your most common mood: {most_common} ({mood_freq[most_common]} times)"

# ---------- HOME ----------
@app.route('/')
def home():
    return render_template('home.html')

# ---------- QR GENERATION ----------
@app.route('/generate', methods=['POST'])
def generate():
    cid = request.form['cid']
    def get_local_ip():
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        try:
            s.connect(('8.8.8.8', 80))
            ip = s.getsockname()[0]
        except Exception:
            ip = '127.0.0.1'
        finally:
            s.close()
        return ip

    local_ip = get_local_ip()
    port = 5000
    if local_ip.startswith('127.'):
        host_base = request.host_url.rstrip('/')
    else:
        host_base = f"http://{local_ip}:{port}"

    link = f"{host_base}/customer/{cid}"
    img = qrcode.make(link)
    if not os.path.exists('static'):
        os.mkdir('static')
    qr_filename = f"qr_{cid}.png"
    qr_path = os.path.join('static', qr_filename)
    img.save(qr_path)
    return render_template('generate.html', link=link, qr_filename=qr_filename)

# ---------- CUSTOMER MOOD - REAL-TIME EMOTION INPUT ----------
@app.route('/customer/<cid>', methods=['GET', 'POST'])
def customer(cid):
    if request.method == 'POST':
        # Check if PDF file was uploaded
        if 'pdf_file' in request.files:
            pdf_file = request.files['pdf_file']
            if pdf_file and pdf_file.filename.endswith('.pdf'):
                # Extract data from PDF
                extracted_data = extract_data_from_pdf(pdf_file)
                
                if extracted_data:
                    # Validate and use extracted data
                    name = extracted_data.get('name', '').strip()
                    age = extracted_data.get('age')
                    mobile = extracted_data.get('mobile', '').strip()
                    email = extracted_data.get('email', '').strip()
                    
                    if not name:
                        return render_template('customer_details.html', cid=cid, error='PDF does not contain valid customer name'), 400
                    
                    # Save PDF file
                    pdf_filename = save_uploaded_pdf(pdf_file, cid)
                    
                    # Check if customer exists
                    existing_customer = Customer.query.filter_by(customer_id=cid).first()
                    if existing_customer:
                        existing_customer.name = name
                        existing_customer.age = age
                        existing_customer.mobile = mobile
                        existing_customer.email = email
                        existing_customer.pdf_filename = pdf_filename
                        existing_customer.pdf_uploaded_at = datetime.now()
                        existing_customer.updated_at = datetime.now()
                    else:
                        customer = Customer(
                            customer_id=cid,
                            name=name,
                            age=age,
                            mobile=mobile,
                            email=email,
                            pdf_filename=pdf_filename,
                            pdf_uploaded_at=datetime.now()
                        )
                        db.session.add(customer)
                    
                    db.session.commit()
                    return redirect(url_for('mood_input', cid=cid))
                else:
                    return render_template('customer_details.html', cid=cid, error='Could not extract data from PDF. Please fill form manually.'), 400
            else:
                return render_template('customer_details.html', cid=cid, error='Please upload a valid PDF file'), 400
        
        # Check if it's a details form or mood form
        elif 'name' in request.form:
            # Handle customer details
            name = request.form.get('name', '').strip()
            age = request.form.get('age', type=int) if request.form.get('age') else None
            mobile = request.form.get('mobile', '').strip()
            email = request.form.get('email', '').strip()
            
            # Validate inputs
            if not name:
                return render_template('customer_details.html', cid=cid, error='Name is required'), 400
            if mobile and not re.match(r'^\d{10}$', mobile.replace('-', '').replace(' ', '')):
                return render_template('customer_details.html', cid=cid, error='Mobile should be 10 digits'), 400
            
            # Check if customer exists
            existing_customer = Customer.query.filter_by(customer_id=cid).first()
            if existing_customer:
                existing_customer.name = name
                existing_customer.age = age
                existing_customer.mobile = mobile
                existing_customer.email = email
                existing_customer.updated_at = datetime.now()
            else:
                customer = Customer(
                    customer_id=cid,
                    name=name,
                    age=age,
                    mobile=mobile,
                    email=email
                )
                db.session.add(customer)
            
            db.session.commit()
            return redirect(url_for('mood_input', cid=cid))
        else:
            # Handle mood form
            mood = request.form.get('mood')
            intensity = request.form.get('intensity', '3')
            notes = request.form.get('notes', '')
            
            if not mood:
                return render_template('customer.html', cid=cid, emotion_levels=EMOTION_LEVELS, error='Please select a mood'), 400
            
            # Ensure customer exists
            customer = Customer.query.filter_by(customer_id=cid).first()
            if not customer:
                return redirect(url_for('customer', cid=cid))
            
            # Store in database
            mood_record = Mood(
                customer_id=cid,
                mood=mood,
                intensity=int(intensity),
                notes=notes
            )
            db.session.add(mood_record)
            
            # Also update customer_data for backwards compatibility
            customer_data[cid] = mood
            
            # Store in history with timestamp for trend analysis
            mood_entry = {
                'mood': mood,
                'intensity': int(intensity),
                'notes': notes,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            user_history[cid].append(mood_entry)
            
            db.session.commit()
            
            # Get suggestion
            suggestion = ai_suggestion(cid, mood)
            
            return render_template('customer_thanks.html', 
                                 mood=mood, 
                                 suggestion=suggestion,
                                 personal_rec=get_personalized_recommendations(cid),
                                 customer_id=cid)

    # Check if customer details exist
    existing_customer = Customer.query.filter_by(customer_id=cid).first()
    if not existing_customer:
        return render_template('customer_details.html', cid=cid)
    
    return render_template('customer.html', cid=cid, emotion_levels=EMOTION_LEVELS)

# ---------- MOOD INPUT PAGE ----------
@app.route('/customer/<cid>/mood', methods=['GET'])
def mood_input(cid):
    """Redirect to mood input after customer details are saved"""
    customer = Customer.query.filter_by(customer_id=cid).first()
    if not customer:
        return redirect(url_for('customer', cid=cid))
    return render_template('customer.html', cid=cid, emotion_levels=EMOTION_LEVELS, customer=customer)

# ---------- REAL-TIME EMOTION TRACKING (API) ----------
@app.route('/api/mood/<cid>', methods=['POST'])
def api_mood(cid):
    """Real-time mood update via mobile/web"""
    data = request.get_json()
    mood = data.get('mood')
    intensity = data.get('intensity', 3)
    
    customer_data[cid] = mood
    mood_entry = {
        'mood': mood,
        'intensity': int(intensity),
        'notes': data.get('notes', ''),
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    user_history[cid].append(mood_entry)
    
    return jsonify({
        'status': 'success',
        'suggestion': ai_suggestion(cid, mood),
        'message': 'Mood recorded successfully'
    })

# ---------- ADMIN LOGIN ----------
@app.route('/admin/login', methods=['GET','POST'])
def admin_login():
    if request.method == 'POST':
        if request.form['user'] == ADMIN_USER and request.form['pass'] == ADMIN_PASS:
            session['admin'] = True
            return redirect('/admin/dashboard')
        return "<h3>Invalid Login</h3>"
    return render_template('admin_login.html')

# ---------- ADMIN VIEW DASHBOARD WITH CUSTOMER DETAILS ----------
@app.route('/admin/dashboard')
def dashboard():
    if not session.get('admin'):
        return redirect('/admin/login')
    
    items = []
    customers = Customer.query.all()
    
    for customer in customers:
        cid = customer.customer_id
        
        # Get latest mood from database
        latest_mood = Mood.query.filter_by(customer_id=cid).order_by(Mood.timestamp.desc()).first()
        mood = latest_mood.mood if latest_mood else 'Not recorded'
        mood_timestamp = latest_mood.timestamp.strftime('%Y-%m-%d %H:%M:%S') if latest_mood else 'Never'
        
        # Also check in-memory data for backwards compatibility
        if mood == 'Not recorded' and cid in customer_data:
            mood = customer_data[cid]
            mood_timestamp = 'Recently'
        
        suggestion = ai_suggestion(cid, mood) if mood != 'Not recorded' else 'No mood recorded yet'
        
        # Check if admin has overridden this suggestion
        if cid in admin_overrides:
            suggestion = admin_overrides[cid]
        
        # Get personal recommendation
        person_rec = get_personalized_recommendations(cid)
        
        # Get mood history count from database
        mood_history_count = Mood.query.filter_by(customer_id=cid).count()
        if cid in user_history:
            mood_history_count += len(user_history[cid])
        
        # Get trend data
        if mood_history_count > 0:
            trend = f"Total moods recorded: {mood_history_count}"
        else:
            trend = "No mood history"
        
        items.append({
            'cid': cid,
            'name': customer.name,
            'age': customer.age,
            'mobile': customer.mobile,
            'email': customer.email,
            'mood': mood,
            'mood_timestamp': mood_timestamp,
            'suggestion': suggestion,
            'personal_rec': person_rec,
            'trend': trend,
            'history_count': mood_history_count,
            'pdf_filename': customer.pdf_filename,
            'pdf_uploaded_at': customer.pdf_uploaded_at.strftime('%Y-%m-%d %H:%M:%S') if customer.pdf_uploaded_at else 'Not uploaded'
        })
    
    return render_template('dashboard.html', items=items)


# ---------- ADMIN RECOMMENDATION OVERRIDE ----------
@app.route('/admin/override/<cid>', methods=['POST'])
def override_suggestion(cid):
    if not session.get('admin'):
        return redirect('/admin/login')
    
    override_text = request.form.get('override_suggestion')
    admin_overrides[cid] = override_text
    
    return redirect('/admin/dashboard')

# ---------- REMOVE OVERRIDE ----------
@app.route('/admin/remove-override/<cid>', methods=['POST'])
def remove_override(cid):
    if not session.get('admin'):
        return redirect('/admin/login')
    
    admin_overrides.pop(cid, None)
    return redirect('/admin/dashboard')

# ---------- ADVANCED ANALYTICS WITH HEATMAP & TRENDS ----------
@app.route('/admin/analytics')
def analytics():
    if not session.get('admin'):
        return redirect('/admin/login')
    
    moods = list(customer_data.values())
    if not moods:
        return "<h3>No data available</h3>"
    
    mood_count = {m: moods.count(m) for m in set(moods)}
    
    if not os.path.exists('static'):
        os.mkdir('static')
    
    # Create mood distribution chart
    plt.figure(figsize=(12, 5))
    
    # Subplot 1: Bar chart
    plt.subplot(1, 2, 1)
    plt.bar(list(mood_count.keys()), list(mood_count.values()), color='steelblue')
    plt.title("Customer Mood Distribution")
    plt.xlabel("Mood")
    plt.ylabel("Count")
    plt.xticks(rotation=45)
    
    # Subplot 2: Pie chart
    plt.subplot(1, 2, 2)
    colors = plt.cm.Pastel1(np.linspace(0, 1, len(mood_count)))
    plt.pie(list(mood_count.values()), labels=list(mood_count.keys()), autopct='%1.1f%%', colors=colors)
    plt.title("Mood Percentage Distribution")
    
    plt.tight_layout()
    mood_path = os.path.join('static', 'mood.png')
    plt.savefig(mood_path, dpi=100)
    plt.close()
    
    return render_template('analytics.html', mood_img='mood.png')

# ---------- MOOD HEATMAP & TRENDS ----------
@app.route('/admin/trends')
def trends():
    if not session.get('admin'):
        return redirect('/admin/login')
    
    if not user_history:
        return "<h3>No historical data available</h3>"
    
    # Analyze trends
    trend_data = {}
    for cid, history in user_history.items():
        if not history:
            continue
        
        mood_count = defaultdict(int)
        for entry in history:
            mood_count[entry['mood']] += 1
        
        trend_data[cid] = dict(mood_count)
    
    # Generate heatmap
    if not os.path.exists('static'):
        os.mkdir('static')
    
    plt.figure(figsize=(14, 6))
    
    # Create trend visualization
    cids = list(trend_data.keys())[:10]  # Top 10 users
    moods_set = set()
    for cid in cids:
        moods_set.update(trend_data[cid].keys())
    moods_list = sorted(list(moods_set))
    
    # Build heatmap data
    heatmap_data = np.zeros((len(cids), len(moods_list)))
    for i, cid in enumerate(cids):
        for j, mood in enumerate(moods_list):
            heatmap_data[i][j] = trend_data[cid].get(mood, 0)
    
    # Plot heatmap
    plt.imshow(heatmap_data, cmap='YlOrRd', aspect='auto')
    plt.colorbar(label='Frequency')
    plt.xticks(range(len(moods_list)), moods_list, rotation=45, ha='right')
    plt.yticks(range(len(cids)), cids)
    plt.title("Mood Heatmap by User")
    plt.xlabel("Mood")
    plt.ylabel("Customer ID")
    plt.tight_layout()
    
    heatmap_path = os.path.join('static', 'heatmap.png')
    plt.savefig(heatmap_path, dpi=100)
    plt.close()
    
    return render_template('trends.html', heatmap_img='heatmap.png', trend_data=trend_data)

# ---------- USER HISTORY PAGE ----------
@app.route('/admin/history/<cid>')
def user_history_page(cid):
    if not session.get('admin'):
        return redirect('/admin/login')
    
    history = user_history.get(cid, [])
    return render_template('user_history.html', cid=cid, history=history, 
                          personal_rec=get_personalized_recommendations(cid))

# ---------- MOBILE QR FEEDBACK LOOP ----------
@app.route('/feedback/<cid>')
def feedback_form(cid):
    """Simplified feedback form for mobile QR scan"""
    current_mood = customer_data.get(cid, 'Unknown')
    suggestion = ai_suggestion(cid, current_mood) if current_mood != 'Unknown' else 'No suggestion yet'
    return render_template('mobile_feedback.html', cid=cid, current_mood=current_mood, suggestion=suggestion)

@app.route('/api/feedback/<cid>', methods=['POST'])
def submit_feedback(cid):
    """Store feedback from mobile QR scan"""
    data = request.get_json()
    feedback = data.get('feedback', '')
    helpful = data.get('helpful', False)
    
    if cid in user_history and len(user_history[cid]) > 0:
        # Add feedback to last entry
        user_history[cid][-1]['feedback'] = feedback
        user_history[cid][-1]['helpful'] = helpful
    
    return jsonify({'status': 'success', 'message': 'Feedback recorded'})

# ---------- DASHBOARD STATS API ----------
@app.route('/api/stats')
def get_stats():
    """Get real-time stats for dashboard"""
    if not session.get('admin'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    total_users = len(customer_data)
    total_entries = sum(len(h) for h in user_history.values())
    
    mood_dist = {}
    for mood in customer_data.values():
        mood_dist[mood] = mood_dist.get(mood, 0) + 1
    
    return jsonify({
        'total_users': total_users,
        'total_entries': total_entries,
        'mood_distribution': mood_dist
    })

# ---------- CUSTOMER FEEDBACK ----------
@app.route('/customer/<cid>/feedback', methods=['GET', 'POST'])
def customer_feedback(cid):
    if request.method == 'POST':
        rating = request.form.get('rating', 5, type=int)
        suggestion = request.form.get('suggestion', '')
        
        # Save feedback to database
        feedback = Feedback(
            customer_id=cid,
            rating=rating,
            suggestion=suggestion,
            timestamp=datetime.now()
        )
        db.session.add(feedback)
        db.session.commit()
        
        return render_template('feedback_success.html', customer_id=cid)
    
    return render_template('customer_feedback.html', cid=cid)

# ---------- ADMIN VIEW FEEDBACKS ----------
@app.route('/admin/feedbacks')
def view_feedbacks():
    if not session.get('admin'):
        return redirect('/admin/login')
    
    # Get all feedbacks
    feedbacks = Feedback.query.all()
    
    # Sort by timestamp descending (latest first)
    feedbacks = sorted(feedbacks, key=lambda x: x.timestamp, reverse=True)
    
    # Count unread feedbacks
    unread_count = Feedback.query.filter_by(status='unread').count()
    
    return render_template('admin_feedbacks.html', feedbacks=feedbacks, unread_count=unread_count)

# ---------- MARK FEEDBACK AS READ ----------
@app.route('/admin/feedback/<int:feedback_id>/mark-read', methods=['POST'])
def mark_feedback_read(feedback_id):
    if not session.get('admin'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    feedback = Feedback.query.get(feedback_id)
    if feedback:
        feedback.status = 'read'
        db.session.commit()
        return jsonify({'status': 'success'})
    return jsonify({'status': 'error'}), 404

# ---------- DELETE FEEDBACK ----------
@app.route('/admin/feedback/<int:feedback_id>/delete', methods=['POST'])
def delete_feedback(feedback_id):
    if not session.get('admin'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    feedback = Feedback.query.get(feedback_id)
    if feedback:
        db.session.delete(feedback)
        db.session.commit()
        return jsonify({'status': 'success'})
    return jsonify({'status': 'error'}), 404

# ---------- CUSTOMER VIEW FOOD MENU ----------
@app.route('/customer/<cid>/food-menu')
def customer_food_menu(cid):
    food_items = FoodItem.query.filter_by(is_available=True).all()
    return render_template('customer_food_menu.html', cid=cid, food_items=food_items)

# ---------- CUSTOMER BUY FOOD ----------
@app.route('/customer/<cid>/buy-food', methods=['POST'])
def customer_buy_food(cid):
    food_item_id = request.form.get('food_item_id', type=int)
    quantity = request.form.get('quantity', 1, type=int)
    
    food_item = FoodItem.query.get(food_item_id)
    
    if not food_item or quantity > food_item.quantity:
        return render_template('purchase_error.html', message='Item not available or insufficient quantity'), 400
    
    # Update inventory
    food_item.quantity -= quantity
    if food_item.quantity == 0:
        food_item.is_available = False
    
    # Create transaction record
    transaction = Transaction(
        customer_id=cid,
        food_item_id=food_item_id,
        quantity_purchased=quantity,
        price_paid=food_item.base_price * quantity
    )
    
    db.session.add(transaction)
    db.session.commit()
    
    total_price = food_item.base_price * quantity
    return render_template('purchase_success.html', 
                         food_name=food_item.name, 
                         quantity=quantity, 
                         total_price=total_price,
                         customer_id=cid)

# ---------- ADMIN FOOD MANAGEMENT ----------
@app.route('/admin/food-management')
def food_management():
    if not session.get('admin'):
        return redirect('/admin/login')
    
    food_items = FoodItem.query.all()
    return render_template('admin_food_management.html', food_items=food_items)

# ---------- ADMIN ADD FOOD ITEM ----------
@app.route('/admin/food/add', methods=['POST'])
def add_food_item():
    if not session.get('admin'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    
    food_item = FoodItem(
        name=data.get('name'),
        base_price=float(data.get('base_price', 0)),
        quantity=int(data.get('quantity', 0)),
        image_url=data.get('image_url', '')
    )
    
    db.session.add(food_item)
    db.session.commit()
    
    return jsonify({
        'status': 'success',
        'id': food_item.id,
        'message': 'Food item added successfully'
    })

# ---------- ADMIN UPDATE FOOD ITEM ----------
@app.route('/admin/food/<int:food_id>/update', methods=['POST'])
def update_food_item(food_id):
    if not session.get('admin'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    food_item = FoodItem.query.get(food_id)
    
    if not food_item:
        return jsonify({'status': 'error', 'message': 'Item not found'}), 404
    
    # Update fields
    if 'name' in data:
        food_item.name = data['name']
    if 'base_price' in data:
        food_item.base_price = float(data['base_price'])
    if 'quantity' in data:
        food_item.quantity = int(data['quantity'])
        food_item.is_available = food_item.quantity > 0
    if 'image_url' in data:
        food_item.image_url = data['image_url']
    
    food_item.updated_at = datetime.now()
    db.session.commit()
    
    return jsonify({
        'status': 'success',
        'message': 'Food item updated successfully'
    })

# ---------- ADMIN DELETE FOOD ITEM ----------
@app.route('/admin/food/<int:food_id>/delete', methods=['POST'])
def delete_food_item(food_id):
    if not session.get('admin'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    food_item = FoodItem.query.get(food_id)
    if not food_item:
        return jsonify({'status': 'error', 'message': 'Item not found'}), 404
    
    db.session.delete(food_item)
    db.session.commit()
    
    return jsonify({'status': 'success', 'message': 'Food item deleted'})

# ---------- ADMIN CSV IMPORT FOR FOOD ITEMS ----------
@app.route('/admin/food/csv-import', methods=['POST'])
def csv_import_food():
    """Import food items from CSV file"""
    if not session.get('admin'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        if 'csv_file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No CSV file provided'}), 400
        
        csv_file = request.files['csv_file']
        
        if not csv_file or csv_file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400
        
        if not csv_file.filename.endswith('.csv'):
            return jsonify({'status': 'error', 'message': 'Please upload a CSV file'}), 400
        
        # Read CSV file
        stream = csv_file.stream.read().decode("UTF-8")
        csv_data = csv.DictReader(stream.split('\n'))
        
        imported_count = 0
        skipped_count = 0
        errors = []
        
        for row in csv_data:
            try:
                if not row or not row.get('name'):
                    skipped_count += 1
                    continue
                
                name = row.get('name', '').strip()
                try:
                    base_price = float(row.get('base_price', 0))
                except (ValueError, TypeError):
                    base_price = 0
                
                try:
                    quantity = int(row.get('quantity', 0))
                except (ValueError, TypeError):
                    quantity = 0
                
                image_url = row.get('image_url', '').strip()
                
                # Check if item already exists
                existing = FoodItem.query.filter_by(name=name).first()
                if existing:
                    # Update existing item
                    existing.base_price = base_price
                    existing.quantity = quantity
                    existing.image_url = image_url
                    existing.is_available = quantity > 0
                    existing.updated_at = datetime.now()
                else:
                    # Create new item
                    new_item = FoodItem(
                        name=name,
                        base_price=base_price,
                        quantity=quantity,
                        image_url=image_url,
                        is_available=quantity > 0
                    )
                    db.session.add(new_item)
                
                imported_count += 1
            except Exception as e:
                skipped_count += 1
                errors.append(f"Row: {name if 'name' in row else 'Unknown'} - {str(e)}")
        
        db.session.commit()
        
        message = f"Successfully imported {imported_count} items"
        if skipped_count > 0:
            message += f" (Skipped {skipped_count} items)"
        if errors:
            message += f". Errors: {', '.join(errors[:3])}"
        
        return jsonify({
            'status': 'success',
            'message': message,
            'imported': imported_count,
            'skipped': skipped_count
        })
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Error importing CSV: {str(e)}'}), 400

# ---------- ADMIN VIEW TRANSACTIONS ----------
@app.route('/admin/transactions')
def view_transactions():
    if not session.get('admin'):
        return redirect('/admin/login')
    
    transactions = Transaction.query.all()
    transactions = sorted(transactions, key=lambda x: x.timestamp, reverse=True)
    
    return render_template('admin_transactions.html', transactions=transactions)

# ---------- GENERATE CUSTOMER PDF REPORT ----------
@app.route('/admin/customer/<cid>/pdf')
def generate_customer_pdf(cid):
    """Generate PDF report with customer details and mood history"""
    if not session.get('admin'):
        return redirect('/admin/login')
    
    customer = Customer.query.filter_by(customer_id=cid).first()
    if not customer:
        return "Customer not found", 404
    
    # Create PDF in memory
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=30,
        alignment=1
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2e5c8a'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Title
    elements.append(Paragraph("Customer Details Report", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Customer Info Table
    elements.append(Paragraph("Customer Information", heading_style))
    customer_info = [
        ['Field', 'Details'],
        ['Customer ID', customer.customer_id],
        ['Name', customer.name],
        ['Age', str(customer.age) if customer.age else 'N/A'],
        ['Mobile', customer.mobile if customer.mobile else 'N/A'],
        ['Email', customer.email if customer.email else 'N/A'],
        ['Created Date', customer.created_at.strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    customer_table = Table(customer_info, colWidths=[2*inch, 4*inch])
    customer_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2e5c8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0f0f0')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')])
    ]))
    elements.append(customer_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Mood History
    elements.append(Paragraph("Mood History", heading_style))
    history = user_history.get(cid, [])
    
    if history:
        history_data = [['Date/Time', 'Mood', 'Intensity', 'Notes']]
        for entry in history[-10:]:  # Last 10 entries
            history_data.append([
                entry.get('timestamp', 'N/A'),
                entry.get('mood', 'N/A'),
                str(entry.get('intensity', 'N/A')),
                entry.get('notes', '')[:50]  # Truncate notes
            ])
        
        history_table = Table(history_data, colWidths=[1.8*inch, 1.5*inch, 1.2*inch, 1.5*inch])
        history_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2e5c8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0f0f0')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')])
        ]))
        elements.append(history_table)
    else:
        elements.append(Paragraph("No mood history recorded yet.", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    pdf_buffer.seek(0)
    
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f"customer_{cid}_{datetime.now().strftime('%Y%m%d')}.pdf"
    )

# ---------- EXPORT ALL CUSTOMERS TO EXCEL ----------
@app.route('/admin/export-excel')
def export_to_excel():
    """Export all customer data to Excel"""
    if not session.get('admin'):
        return redirect('/admin/login')
    
    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Define styles
    header_fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Customer ID', 'Name', 'Age', 'Mobile', 'Email', 'Current Mood', 'Total Entries', 'Created Date']
    ws.append(headers)
    
    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add customer data
    customers = Customer.query.all()
    for customer in customers:
        cid = customer.customer_id
        mood = customer_data.get(cid, 'Not recorded')
        entries = len(user_history.get(cid, []))
        
        ws.append([
            cid,
            customer.name,
            customer.age if customer.age else '',
            customer.mobile if customer.mobile else '',
            customer.email if customer.email else '',
            mood,
            entries,
            customer.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    # Auto-adjust column widths
    column_widths = [15, 20, 8, 15, 20, 15, 12, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+i)].width = width
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"customers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

# ---------- IMPORT CUSTOMERS FROM EXCEL ----------
@app.route('/admin/import-excel', methods=['GET', 'POST'])
def import_from_excel():
    """Import customer data from Excel file"""
    if not session.get('admin'):
        return redirect('/admin/login')
    
    if request.method == 'POST':
        # Check if file was uploaded
        if 'file' not in request.files:
            return render_template('admin_import.html', error='No file selected'), 400
        
        file = request.files['file']
        if file.filename == '':
            return render_template('admin_import.html', error='No file selected'), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return render_template('admin_import.html', error='Please upload an Excel file'), 400
        
        try:
            # Read Excel file
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            imported_count = 0
            errors = []
            
            # Skip header row
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    cid, name, age, mobile, email = row[0], row[1], row[2], row[3], row[4]
                    
                    if not cid or not name:
                        errors.append(f"Row {row_idx}: Customer ID and Name are required")
                        continue
                    
                    # Validate mobile format if provided
                    if mobile and not re.match(r'^\d{10}$', str(mobile).replace('-', '').replace(' ', '')):
                        errors.append(f"Row {row_idx}: Invalid mobile format (should be 10 digits)")
                        continue
                    
                    # Check if customer exists
                    existing = Customer.query.filter_by(customer_id=str(cid)).first()
                    if existing:
                        existing.name = name
                        existing.age = int(age) if age else None
                        existing.mobile = str(mobile) if mobile else None
                        existing.email = email
                        existing.updated_at = datetime.now()
                    else:
                        customer = Customer(
                            customer_id=str(cid),
                            name=name,
                            age=int(age) if age else None,
                            mobile=str(mobile) if mobile else None,
                            email=email
                        )
                        db.session.add(customer)
                    
                    imported_count += 1
                
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
            
            db.session.commit()
            return render_template('admin_import.html', 
                                 success=f'Successfully imported {imported_count} customers',
                                 errors=errors)
        
        except Exception as e:
            return render_template('admin_import.html', error=f'Error processing file: {str(e)}'), 400
    
    return render_template('admin_import.html')

# ---------- API: GET FOOD ITEMS ----------
@app.route('/api/food-items')
def api_food_items():
    food_items = FoodItem.query.filter_by(is_available=True).all()
    items_data = [{
        'id': item.id,
        'name': item.name,
        'base_price': item.base_price,
        'quantity': item.quantity,
        'image_url': item.image_url
    } for item in food_items]
    return jsonify(items_data)

# ---------- LOGOUT ----------
@app.route('/admin/logout')
def logout():
    session.pop('admin', None)
    return redirect('/')

# ---------- DOWNLOAD CUSTOMER PDF ----------
@app.route('/admin/download-pdf/<cid>')
def download_pdf(cid):
    """Download uploaded customer PDF"""
    if not session.get('admin'):
        return redirect('/admin/login')
    
    customer = Customer.query.filter_by(customer_id=cid).first()
    if not customer or not customer.pdf_filename:
        return "PDF not found", 404
    
    pdf_path = os.path.join('static/uploads', customer.pdf_filename)
    if not os.path.exists(pdf_path):
        return "PDF file not found on server", 404
    
    return send_file(
        pdf_path,
        as_attachment=True,
        download_name=customer.pdf_filename
    )

# ---------- ADMIN UPDATE CUSTOMER PDF ----------
@app.route('/admin/customer/<cid>/update-pdf', methods=['POST'])
def update_customer_pdf(cid):
    """Admin can update/replace customer's PDF"""
    if not session.get('admin'):
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403
    
    customer = Customer.query.filter_by(customer_id=cid).first()
    if not customer:
        return jsonify({'status': 'error', 'message': 'Customer not found'}), 404
    
    if 'pdf_file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file provided'}), 400
    
    pdf_file = request.files['pdf_file']
    if not pdf_file or not pdf_file.filename.endswith('.pdf'):
        return jsonify({'status': 'error', 'message': 'Invalid PDF file'}), 400
    
    try:
        # Delete old PDF if exists
        if customer.pdf_filename:
            old_path = os.path.join('static/uploads', customer.pdf_filename)
            if os.path.exists(old_path):
                os.remove(old_path)
        
        # Save new PDF
        new_filename = save_uploaded_pdf(pdf_file, cid)
        if not new_filename:
            return jsonify({'status': 'error', 'message': 'Failed to save PDF'}), 500
        
        # Update customer record
        customer.pdf_filename = new_filename
        customer.pdf_uploaded_at = datetime.now()
        customer.updated_at = datetime.now()
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'PDF updated successfully',
            'filename': new_filename,
            'upload_date': customer.pdf_uploaded_at.strftime('%Y-%m-%d %H:%M:%S')
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ---------- RUN ----------
if __name__ == '__main__':
    if not os.path.exists("static"):
        os.mkdir("static")
    if not os.path.exists("templates"):
        os.mkdir("templates")
    if not os.path.exists("static/uploads"):
        os.makedirs("static/uploads")
    
    # Create database tables
    with app.app_context():
        db.drop_all()  # Force drop and recreate
        db.create_all()
        print("✅ Database initialized with fresh schema")
    
    # Run on all interfaces so phones on the same network can access via LAN IP
    try:
        app.run(host='0.0.0.0', port=5000, debug=False)  # Changed to False to avoid debug reload issues
    except Exception as e:
        print(f"Error starting Flask app: {e}")
        print("Make sure port 5000 is available")
